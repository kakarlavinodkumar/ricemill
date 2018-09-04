from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.styles import colors
from openpyxl.styles import *
from styles2 import applyStyles
import pyexcel as p
from math import *
import os
from xls_to_xlsx import xls_to_xlsx_converter

def processExcel(filename) :

    xls_to_xlsx_converter(filename);

    workbook = load_workbook(filename+'x')
    work_sheets= workbook.worksheets
    current_worksheet = work_sheets[0]


    wb = Workbook()
    splited = filename.split(".")
    dest_filename = splited[0]+"output"+".xlsx"
    ws1 = wb.active
    ws1.title = "Trading"

    paddy_purchase_flag = False
    paddy_purchases = current_worksheet["D12"].value + current_worksheet["D14"].value + current_worksheet["D15"].value + current_worksheet["D17"].value
    for row in current_worksheet :
        qt_flag = False
        amt_flag= False
        qt_flag1 = False
        amt_flg1 = False 
        new_row = []
        amount = None
        quantity = None
        amount1 = None
        quantity1 = None
        range_limits = [None,None]
        for cell in row :
            ##print  cell.row
            if len(new_row) < 8 :
                new_row.append(cell.value)
            if cell.col_idx == 7 :
                cell_value = cell.value
                if isinstance(cell_value,(int,float,long)) : 
                    quantity = cell.value
                    qt_flag  = True
                    ##print"Qt Flag: True"
            if cell.col_idx == 8 :
                cell_value = cell.value
                if isinstance(cell_value,(int,float,long)) :
                    amount = cell.value
                    amt_flag = True
                    ##print"Amount Flag True"

            if cell.col_idx == 3 :
                cell_value = cell.value
                if isinstance(cell_value,(int,float,long)) :
                    quantity1 = cell.value
                    qt_flag1  = True 
            
            if cell.col_idx == 4 :
                cell_value = cell.value
                if isinstance(cell_value,(int,float,long)) :
                    amount1 = cell.value
                    amt_flag1 = True
            if cell.col_idx == 11 :
                #print"Range   " + str(cell.value)
                #print"Range Type  " + str(type(cell.value))

                if cell.value is not None :
                    range_str = str(cell.value)
                    range_limits = range_str.split("-")
                    range_limits[0] = int(range_limits[0])
                    range_limits[1] = int(range_limits[1])

        new_row.append('')
        if qt_flag1 and amt_flag1 and amount1>0 and quantity1 > 0:
             ##print"Both Flags True  "
            avg1 = amount1 /quantity1
            new_row.append(avg1)
            ##printavg
            lower_limit = 1500
            upper_limit = 2000
            if cell.row <20 :
                if isinstance(range_limits[0],int) and isinstance(range_limits[1],int) :
                    lower_limit = (range_limits[0])
                    upper_limit = (range_limits[1])
                    print "First Limit Ranges:  "+str(range_limits[0])+"   " + str(range_limits[1])

            if avg1 <lower_limit or avg1 >upper_limit :
                new_row.append("Error")
            else :
                new_row.append("Correct")
            qt_flag1 = False
            amt_flag1= False    
        else :
            new_row.append('')
            new_row.append('')

        if qt_flag and amt_flag  and amount > 0 and quantity > 0:
            ##print"Both Flags True  "
            avg = amount /quantity
            ##printavg
            #new_row.append('')
            new_row.append(avg)
            lower_limit = 1500
            upper_limit = 2000
            if cell.row >20 :
                if isinstance(range_limits[0],int) and isinstance(range_limits[1],int) :
                    lower_limit = (range_limits[0])
                    upper_limit = (range_limits[1])
                    print "Second Limit Ranges:  "+str(range_limits[0])+"   " + str(range_limits[1])

            if avg <lower_limit or avg >upper_limit :
                new_row.append("Error")
            else :
                new_row.append("Correct")
            qt_flag = False
            amt_flag= False    
            
        ws1.append(new_row)
    #print"Paddy Purchases  "+str(paddy_purchases)

    #Secong Sheet


    current_worksheet = work_sheets[1]
    paddy_milling = None
    paddy_milling_cell = current_worksheet["F15"]
    paddy_milling = paddy_milling_cell.value
    #print"Paddy Milling  "+str(paddy_milling)



    ws2= wb.create_sheet("Yield")
    for row in current_worksheet : 
        new_row =[]
        for cell in row :
            new_row.append(cell.value)

            if cell.row == 47 and cell.col_idx ==6 :
                if cell.value > 69 and cell.value <70 :
                    new_row.append("Correct")
                    #print"Yield Error"
                else :
                    #print"Yield Correct"
                    new_row.append("Error")
            # if "Yield=" in new_row :
            #     if isinstance(cell.value,(int,long,float)) :
            #         value = cell.value * 100
            #         if value > 69 and value < 70 :
            #             new_row.append("Correct")
            #         else :
            #             new_row.append("Error")
        ws2.append(new_row)


    #Third Sheet


    current_worksheet = work_sheets[2]
    electricity_charges = None
    electricity_charges_cell = current_worksheet["C11"]
    electricity_charges = electricity_charges_cell.value


    #Hamali Charges
    hamali_charges = None
    hamali_charges = current_worksheet["C28"].value
    hamali_charges_flag = None
    #print"Hamali Charges  "+ str(hamali_charges)
    #print"Paddy Milling 35%  ",(int(paddy_purchases)* 35)/100
    if hamali_charges >( paddy_purchases * 35)/100 :
        hamali_charges_flag = True
        #print"Hamali Charges Error"
    else :
        hamali_charges_flag = False
        #print"Hamali Charges Correct"





    #Electricity Charges
    #print"Electricity Charges  "+str(electricity_charges)
    #print"Paddy Milling * 60   " +str(paddy_milling*60)
    #print"Paddy Milling *75  " +str(paddy_milling*75)
    electricity_flag= None
    if paddy_milling * 60 < electricity_charges and paddy_milling*75 > electricity_charges :
        electricity_flag = True
        #print"Correct"        
    else :
        electricity_flag = False
        #print"Error"


    #Labour Charges

    labour_flag = None
    labour_charges = current_worksheet["C29"].value
    if (paddy_purchases *4)/100 > labour_charges :
        labour_flag = True
    else :
        labour_flag = False

    ws3= wb.create_sheet("Manu&P&L")
    for row in current_worksheet : 
        new_row =[]
        for cell in row :
            if cell.row == 11 and cell.col_idx == 4:
                if electricity_flag :
                    new_row.append("Correct")
                else :
                    new_row.append("Error")
            elif cell.row == 28 and cell.col_idx == 4:
                if hamali_charges_flag :
                    new_row.append("Error")
                else :
                    new_row.append("Correct")

            elif cell.row == 29 and cell.col_idx == 4:
                if labour_flag :
                    new_row.append("Error")
                else :
                    new_row.append("Correct")
            else :
                new_row.append(cell.value)

            # if "Yield" in new_row :
            #     if isinstance(cell.value,(int,long,float)) :
            #         value = cell.value * 100
            #         if value > 69 and value < 70 :
            #             new_row.append("Correct")
            #         else :
            #             new_row.append("Error")
        ws3.append(new_row)

    #Fourth Sheet

    current_worksheet = work_sheets[3]
    ws4= wb.create_sheet("B.S")

    prev_cell_value = None
    trade_creditors = None
    sunday_debtors = None


    for row in current_worksheet : 
        new_row =[]
        for cell in row :
            
            if cell.row == 24 and cell.col_idx == 3 :
                if paddy_milling /4 > trade_creditors :
                    new_row.append("Error")
                else :
                    new_row.append("Correct")    
            elif cell.row == 24 and cell.col_idx == 5:
                new_row.append(cell.value)
                if cell.value > (paddy_purchases*15)/100 :
                    new_row.append("Error")
                else :
                    new_row.append("Correct")     
            else :
                new_row.append(cell.value)    
            if prev_cell_value == "Trade Creditors" :
                trade_creditors = cell.value
            if prev_cell_value == "Sundry Debtors" :
                sunday_debtors = cell.value
            # if "Yield" in new_row :
            #     if isinstance(cell.value,(int,long,float)) :
            #         value = cell.value * 100
            #         if value > 69 and value < 70 :
            #             new_row.append("Correct")
            #         else :
            #             new_row.append("Error")
            prev_cell_value = cell.value


        ws4.append(new_row)

    #print"Trade Creditors " +str(trade_creditors)
    #print"Sunday Debtors  "+str(sunday_debtors)

    #Fifth Sheet

    current_worksheet = work_sheets[4]
    ws5= wb.create_sheet("Capital ac")
    for row in current_worksheet : 
        new_row =[]
        for cell in row :
            new_row.append(cell.value)
            # if "Yield" in new_row :
            #     if isinstance(cell.value,(int,long,float)) :
            #         value = cell.value * 100
            #         if value > 69 and value < 70 :
            #             new_row.append("Correct")
            #         else :
            #             new_row.append("Error")
        ws5.append(new_row)


    #Sixth Sheet

    current_worksheet = work_sheets[5]
    ws6= wb.create_sheet("Dep.Schedule")
    for row in current_worksheet : 
        new_row =[]
        for cell in row :
            new_row.append(cell.value)
            # if "Yield" in new_row :
            #     if isinstance(cell.value,(int,long,float)) :
            #         value = cell.value * 100
            #         if value > 69 and value < 70 :
            #             new_row.append("Correct")
            #         else :
            #             new_row.append("Error")
        ws6.append(new_row)


    #Seventh Sheet


    current_worksheet = work_sheets[6]
    ws7= wb.create_sheet("BS WORKING")
    for row in current_worksheet : 
        new_row =[]
        for cell in row :
            new_row.append(cell.value)
            # if "Yield" in new_row :
            #     if isinstance(cell.value,(int,long,float)) :
            #         value = cell.value * 100
            #         if value > 69 and value < 70 :
            #             new_row.append("Correct")
            #         else :
            #             new_row.append("Error")
        ws7.append(new_row)


    #Eight Sheet


    current_worksheet = work_sheets[7 ]
    ws2= wb.create_sheet("Analysis")
    for row in current_worksheet : 
        new_row =[]
        for cell in row :
            new_row.append(cell.value)
            # if "Yield" in new_row :
            #     if isinstance(cell.value,(int,long,float)) :
            #         value = cell.value * 100
            #         if value > 69 and value < 70 :
            #             new_row.append("Correct")
            #         else :
            #             new_row.append("Error")
        ws2.append(new_row)

    wb.save(filename = dest_filename)





    #Apply Styles To First Sheet

    applyStyles()